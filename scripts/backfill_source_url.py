"""One-shot backfill: record which page each stored contact was read from.

``source_url`` is new, so every record written before it is blank. The parser
now stamps it (see ``parse_directory``) and the coaches-database import takes
the workbook's "Landing pages" column, but neither helps rows already stored:
re-scraping only reaches sites that answer, and the imported rows have no site
at all.

Two sources fill them in, both already on disk:

* **The run reports.** ``data/runs/*/sites.json`` records the directory URL
  actually fetched for each host. That is the same page the contacts from that
  host were parsed out of, so the join is exact rather than a reconstruction.
  Later runs win, since a site can move its directory.
* **The coaches database.** Its "Landing pages" column is the team page each
  imported row came from, matched on school name and state the same way the
  import matched them in the first place.

Anything neither source covers is left blank and reported.

    python scripts/backfill_source_url.py            # report
    python scripts/backfill_source_url.py --apply    # rewrite, keeping a .bak
"""

from __future__ import annotations

import argparse
import collections
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scrapbot import importer  # noqa: E402
from scrapbot.config import Settings  # noqa: E402
from scrapbot.storage import ContactStore, StoreLock  # noqa: E402

WORKBOOK = "Men_s_Basketball_Coaches_Database__April_2026_.xlsx"
SHEETS = ["DI", "DII", "DIII", "JuCo", "NAIA"]


def urls_from_runs(settings: Settings) -> dict[str, str]:
    """``host -> directory URL``, newest run wins."""
    out: dict[str, str] = {}
    for report in sorted((settings.data_dir / "runs").glob("*/sites.json")):
        try:
            rows = json.loads(report.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            continue
        for row in rows if isinstance(rows, list) else rows.get("sites", []):
            host, url = row.get("domain"), row.get("url")
            # Only a run that found people knows the page they came from; a
            # failure's url is whatever was tried last.
            #
            # A manual run records the saved file it parsed rather than a URL.
            # That is still where the row came from — and for a blocked host it
            # is the only provenance there is — so it is kept as written.
            if host and url and row.get("people"):
                out[host] = url
    return out


def urls_from_workbook(path: Path) -> dict[tuple[str, str], str]:
    """``(normalised school, state) -> team page`` from the coaches database."""
    out: dict[tuple[str, str], str] = {}
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in SHEETS:
        started = False
        for raw in wb[sheet].iter_rows(values_only=True):
            if not started:
                started = bool(raw) and str(raw[0] or "").strip() == "Conference"
                continue
            if not raw or len(raw) < 14:
                continue
            school, state, landing = raw[7], raw[1], raw[13]
            if not school or not landing or not str(landing).startswith("http"):
                continue
            key = (importer.normalize_name(str(school)), str(state or "").strip().lower())
            out.setdefault(key, str(landing).strip())
    wb.close()
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="Rewrite the store.")
    parser.add_argument("--data-dir", type=Path)
    parser.add_argument("--workbook", type=Path, default=Path(WORKBOOK))
    args = parser.parse_args()

    settings = Settings()
    if args.data_dir:
        settings.data_dir = args.data_dir

    by_host = urls_from_runs(settings)
    by_school = urls_from_workbook(args.workbook)
    print(f"{len(by_host)} host(s) with a directory URL in the run reports")
    print(f"{len(by_school)} school(s) with a landing page in the workbook")

    store = ContactStore(settings).load()
    contacts = list(store.leads.values())

    from_runs = from_book = already = missing = 0
    unresolved: collections.Counter = collections.Counter()

    # The workbook is keyed on school + state; a contact carries no state, so
    # match on the name and only where every sheet agrees on one link.
    book_by_name: dict[str, set[str]] = collections.defaultdict(set)
    for (name, _state), url in by_school.items():
        book_by_name[name].add(url)

    for contact in contacts:
        if contact.source_url:
            already += 1
            continue

        # The workbook's "Landing pages" is the men's basketball page itself —
        # https://aamusports.com/sports/mens-basketball — which is a more exact
        # answer to "where did this come from" than the staff directory the
        # scraper read. So it wins, but *only for men's basketball staff*: it
        # is one link per school, and handing it to a softball coach would
        # state something untrue about where their row came from.
        book = book_by_name.get(importer.normalize_name(contact.school or "")) or set()
        # next(iter(...)), not pop(): the set is shared by every contact at
        # this school, and popping it empties it for all but the first.
        exact = next(iter(book)) if len(book) == 1 else None
        basketball = "men's basketball" in (contact.sport or "").lower()

        url = None
        if exact and basketball:
            url, from_book = exact, from_book + 1
        elif by_host.get(contact.school_domain):
            url, from_runs = by_host[contact.school_domain], from_runs + 1
        elif exact:
            # Nothing better on record — the school's own team page still says
            # more about the source than nothing at all.
            url, from_book = exact, from_book + 1

        if url:
            if args.apply:
                contact.source_url = url
        else:
            missing += 1
            unresolved[contact.school_domain] += 1

    print(f"\n{len(contacts)} contact(s) in {settings.contacts_path}")
    print(f"  already had one          : {already}")
    print(f"  filled from run reports  : {from_runs}")
    print(f"  filled from the workbook : {from_book}")
    print(f"  still blank              : {missing}")

    if unresolved:
        print(f"\n  no page on record for {len(unresolved)} host(s):")
        for host, count in unresolved.most_common(10):
            print(f"    {host:<32} {count}")

    if not args.apply:
        print("\nreport only — nothing written. Re-run with --apply.")
        return 0
    if not (from_runs or from_book):
        print("\nnothing to change")
        return 0

    backup = settings.contacts_path.with_suffix(".json.bak")
    n = 2
    while backup.exists():
        backup = settings.contacts_path.with_suffix(f".json.bak{n}")
        n += 1
    shutil.copy2(settings.contacts_path, backup)
    print(f"\nbacked up to {backup}")
    with StoreLock(settings):
        store.save()
    print(f"filled {from_runs + from_book} source URL(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
