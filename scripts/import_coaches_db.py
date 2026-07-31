"""Import men's basketball coaches from the supplied coaches database.

Records are tagged ``source = "coaches-db"``, which is how they are told apart
from the scraped ones in the same store.

Only schools whose athletics host is *blocked* are imported. Everywhere else
the scraper can read the staff directory itself, which yields the whole
department rather than the basketball staff alone.

    python scripts/import_coaches_db.py                     # report only
    python scripts/import_coaches_db.py --apply             # write the store
    python scripts/import_coaches_db.py --blocked FILE      # a different host list
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scrapbot import extract, importer, storage  # noqa: E402
from scrapbot.config import Settings  # noqa: E402
from scrapbot.models import Contact  # noqa: E402
from scrapbot.sources.coaches import is_coaching_title, normalize_person_name  # noqa: E402

WORKBOOK = "Men_s_Basketball_Coaches_Database__April_2026_.xlsx"
SHEETS = ["DI", "DII", "DIII", "JuCo", "NAIA"]
SOURCE = "coaches-db"

# Column positions, identical on all five sheets. The header sits on row 6
# (1-based); rows 1-5 are the copyright notice and the column-group banners.
HEADER_MARKER = "Conference"
COL = {
    "state": 1,
    "removed": 6,
    "school": 7,
    "first": 8,
    "last": 9,
    "position": 10,
    "email": 11,
    "phone": 12,
}


def _clean(value) -> str:
    """Cell text, with the sheet's own empty-marker treated as empty."""
    text = str(value).strip() if value is not None else ""
    return "" if text in ("-", "--", "n/a", "N/A") else text


def read_blocked(path: Path) -> dict[tuple[str, str], str]:
    """``(normalised school, state) -> athletics host`` from a seed file.

    The seed file's comment carries the school and state, which is what the
    coaches database can be matched on -- it has no hostnames of its own.
    """
    out: dict[tuple[str, str], str] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        host, _, comment = line.partition("#")
        host = host.strip()
        comment = comment.strip()
        if not host or "(" not in comment:
            continue
        school, _, state = comment.rpartition("(")
        key = (importer.normalize_name(school.strip()), state.rstrip(")").strip().lower())
        out.setdefault(key, host)
    return out


def rows(workbook: Path):
    """Yield ``(sheet, row)`` for every data row across the five sheets."""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    for name in SHEETS:
        ws = wb[name]
        started = False
        for raw in ws.iter_rows(values_only=True):
            if not started:
                started = bool(raw) and str(raw[0] or "").strip() == HEADER_MARKER
                continue
            if raw and any(v is not None for v in raw):
                yield name, raw
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="Write to the contact store.")
    parser.add_argument(
        "--blocked",
        type=Path,
        default=Path("data/seeds/blocked-from-list.txt"),
        help="Seed file of blocked hosts, with '# School (State)' comments.",
    )
    parser.add_argument("--workbook", type=Path, default=Path(WORKBOOK))
    parser.add_argument("--data-dir", type=Path)
    args = parser.parse_args()

    settings = Settings()
    if args.data_dir:
        settings.data_dir = args.data_dir

    blocked = read_blocked(args.blocked)
    print(f"{len(blocked)} blocked school(s) from {args.blocked}")

    stats = Counter()
    contacts: list[Contact] = []
    per_school: dict[str, int] = defaultdict(int)
    unmatched: Counter = Counter()

    for sheet, raw in rows(args.workbook):
        def cell(field: str) -> str:
            index = COL[field]
            return _clean(raw[index]) if index < len(raw) else ""

        stats["rows"] += 1
        if cell("removed").lower() == "y":
            stats["removed"] += 1
            continue

        first, last = cell("first"), cell("last")
        if not first or not last:
            # Removed-person rows put a prose note where the name goes.
            stats["no name"] += 1
            continue

        school, state = cell("school"), cell("state")
        host = blocked.get((importer.normalize_name(school), state.lower()))
        if not host:
            stats["not a blocked school"] += 1
            unmatched[f"{school} ({state})"] += 1
            continue

        title = cell("position") or None
        contact = Contact(
            name=normalize_person_name(f"{first} {last}"),
            school_domain=host,
            school=school or None,
            title=title,
            sport="Men's Basketball",
            # profile_url stays empty on purpose. The sheet's "Landing pages"
            # is the *team* page, shared by every coach at the school, and the
            # store keys on profile_url first -- using it would collapse a
            # whole staff into a single record.
            is_coach=is_coaching_title(title),
            source=SOURCE,
        )
        email = cell("email").lower()
        if email and extract.EMAIL_RE.fullmatch(email) and not extract.EMAIL_NOISE.search(email):
            contact.emails.append(email)
        else:
            stats["no usable email"] += 1
        phone = extract.clean_phone(cell("phone"))
        if phone:
            contact.phones.append(phone)

        contacts.append(contact)
        per_school[host] += 1
        stats["importable"] += 1

    print()
    for label in (
        "rows", "removed", "no name", "not a blocked school", "importable", "no usable email"
    ):
        print(f"  {label:<24} {stats[label]}")
    print(f"  {'schools covered':<24} {len(per_school)}")
    print(f"  {'with an email':<24} {sum(1 for c in contacts if c.emails)}")
    print(f"  {'with a phone':<24} {sum(1 for c in contacts if c.phones)}")
    print(f"  {'coaching titles':<24} {sum(1 for c in contacts if c.is_coach)}")

    if not args.apply:
        print("\nreport only — nothing written. Re-run with --apply to write.")
        return 0

    with storage.StoreLock(settings):
        store = storage.ContactStore(settings).load()
        before = len(store.leads)
        for contact in contacts:
            store.upsert(contact)
        store.save()
        print(
            f"\ncontacts {before} -> {len(store.leads)} "
            f"({store.new_count} new, {store.updated_count} merged) -> {settings.contacts_path}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
